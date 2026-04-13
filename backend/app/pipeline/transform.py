"""
transform.py — memory-efficient xlsx streaming + schema unification.

Exports:
    stream_workbook_chunks(path, chunk_size) -> Iterator[pd.DataFrame]
    normalize_chunk(raw_df, category)        -> pd.DataFrame

Unified schema (7 columns, `description` carried for table preview only):
    date, product, category, country, value, quantity, description
"""
from __future__ import annotations

from pathlib import Path
from typing import Iterable, Iterator, Optional

import pandas as pd
from openpyxl import load_workbook

from .utils import (
    clean_string,
    normalize_column_name,
    safe_float,
    setup_logger,
)

logger = setup_logger()

UNIFIED_COLUMNS = ["date", "product", "category", "country", "value", "quantity"]

# Each unified key maps to a set of normalized candidate column names.
# Entries must be in normalized form (see normalize_column_name).
COLUMN_ALIASES: dict[str, set[str]] = {
    "date": {
        "date", "shipment_date", "txn_date", "transaction_date",
        "invoice_date", "arrival_date",
    },
    "hs_code": {"hs_code", "hscode", "h_s_code", "hs"},
    "description": {
        "product_description", "product", "description",
        "item_description", "hs_product_description",
    },
    "country": {
        "country_of_origin", "origin_country", "country",
        "country_of_destination", "destination_country",
    },
    "quantity": {
        "standard_qty", "std_quantity", "standard_quantity",
        "quantity", "qty",
    },
    "value": {
        "estimated_cif_value", "cif_value", "value",
        "total_value", "value_usd", "estimated_value",
    },
    "unit_rate": {
        "standard_unit_rate", "std_unit_rate",
        "unit_rate", "unit_rate_usd",
    },
}

_PREFERRED_DATA_SHEETS = ("Data Sheet", "Global Search Data", "Sheet1")
_SKIP_SHEETS = {"CalculationSheet", "ReadMe", "Index", "Cover"}


# ---------------------------------------------------------------------------
# Workbook streaming
# ---------------------------------------------------------------------------

def _find_data_sheet(sheetnames: Iterable[str]) -> Optional[str]:
    names = list(sheetnames)
    for candidate in _PREFERRED_DATA_SHEETS:
        if candidate in names:
            return candidate
    for name in names:
        if name not in _SKIP_SHEETS:
            return name
    return names[0] if names else None


def _looks_like_header_row(row: tuple) -> bool:
    if not row:
        return False
    for cell in row:
        if cell is None:
            continue
        if str(cell).strip().lower() == "date":
            return True
    return False


def stream_workbook_chunks(
    path: Path, chunk_size: int = 5_000,
) -> Iterator[pd.DataFrame]:
    """Yield chunks of raw rows from a Volza workbook.

    openpyxl read-only mode — the full sheet never materializes in memory.
    """
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        sheet_name = _find_data_sheet(wb.sheetnames)
        if sheet_name is None:
            logger.warning("  %s: no usable sheet found", path.name)
            return

        logger.info("  -> sheet: %s", sheet_name)
        ws = wb[sheet_name]
        iterator = ws.iter_rows(values_only=True)

        # Volza files often have 1 decorative row above the real header.
        header: Optional[list[str]] = None
        for i, row in enumerate(iterator):
            if i > 15:
                break
            if _looks_like_header_row(row):
                header = [
                    (str(cell).strip() if cell is not None else f"col_{j}")
                    for j, cell in enumerate(row)
                ]
                break

        if not header:
            logger.warning("  %s: header row not found", path.name)
            return

        n_cols = len(header)
        buffer: list[list] = []
        rows_yielded = 0

        for row in iterator:
            if row is None:
                continue
            if all(cell is None or cell == "" for cell in row):
                continue

            row_list = list(row)
            if len(row_list) < n_cols:
                row_list.extend([None] * (n_cols - len(row_list)))
            elif len(row_list) > n_cols:
                row_list = row_list[:n_cols]

            buffer.append(row_list)
            if len(buffer) >= chunk_size:
                yield pd.DataFrame(buffer, columns=header)
                rows_yielded += len(buffer)
                buffer = []

        if buffer:
            yield pd.DataFrame(buffer, columns=header)
            rows_yielded += len(buffer)

        logger.info("  -> streamed %d rows from %s", rows_yielded, path.name)
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Chunk normalization
# ---------------------------------------------------------------------------

def _build_column_map(raw_columns: Iterable[str]) -> dict[str, str]:
    """Map raw column names to unified schema keys; first-match wins."""
    resolved: dict[str, str] = {}
    for raw in raw_columns:
        norm = normalize_column_name(raw)
        if not norm:
            continue
        for unified_key, aliases in COLUMN_ALIASES.items():
            if unified_key in resolved:
                continue
            if norm in aliases:
                resolved[unified_key] = raw
                break
    return resolved


def _derive_product_key(hs_code: Optional[str], description: Optional[str]) -> str:
    """Prefer HS4 digit prefix; fall back to first 3 words of description."""
    if hs_code:
        digits = "".join(ch for ch in str(hs_code) if ch.isdigit())
        if len(digits) >= 4:
            return digits[:4]
    if description:
        words = str(description).split()
        key = " ".join(words[:3])[:40].upper()
        if key:
            return key
    return "Unknown"


def normalize_chunk(raw_df: pd.DataFrame, category: str) -> pd.DataFrame:
    """Project a raw chunk into the unified schema and clean values."""
    if raw_df.empty:
        return pd.DataFrame(columns=UNIFIED_COLUMNS + ["description"])

    col_map = _build_column_map(raw_df.columns)

    if "date" in col_map:
        dates = pd.to_datetime(raw_df[col_map["date"]], errors="coerce")
    else:
        dates = pd.Series([pd.NaT] * len(raw_df))

    desc_raw = (
        raw_df[col_map["description"]] if "description" in col_map
        else pd.Series([None] * len(raw_df))
    )
    hs_raw = (
        raw_df[col_map["hs_code"]] if "hs_code" in col_map
        else pd.Series([None] * len(raw_df))
    )

    country_raw = (
        raw_df[col_map["country"]] if "country" in col_map
        else pd.Series([None] * len(raw_df))
    )

    if "value" in col_map:
        value_series = raw_df[col_map["value"]].map(safe_float)
    elif "unit_rate" in col_map and "quantity" in col_map:
        qty_tmp = raw_df[col_map["quantity"]].map(safe_float)
        rate_tmp = raw_df[col_map["unit_rate"]].map(safe_float)
        value_series = qty_tmp * rate_tmp
    else:
        value_series = pd.Series([0.0] * len(raw_df))

    if "quantity" in col_map:
        qty_series = raw_df[col_map["quantity"]].map(safe_float)
    else:
        qty_series = pd.Series([0.0] * len(raw_df))

    out = pd.DataFrame({
        "date": dates.dt.strftime("%Y-%m-%d"),
        "product": [
            _derive_product_key(hs, desc) for hs, desc in zip(hs_raw, desc_raw)
        ],
        "category": category,
        "country": [clean_string(c) or "Unknown" for c in country_raw],
        "value": value_series.astype(float).fillna(0.0),
        "quantity": qty_series.astype(float).fillna(0.0),
        "description": [clean_string(d, max_len=120) for d in desc_raw],
    })

    mask_keep = ~(
        (out["value"] == 0)
        & (out["quantity"] == 0)
        & (out["date"].isna())
    )
    out = out[mask_keep].reset_index(drop=True)
    return out
